from doctest import REPORT_ONLY_FIRST_FAILURE
from xml.dom.minidom import Document
import openpyxl as xl
from openpyxl import load_workbook
import streamlit as st

import os
import sys
import urllib.request

from docx.shared import Cm
from docxtpl import DocxTemplate
from docx2pdf import convert

st.title('Meu Labs Report Generation')

report_number = st.number_input('Enter report number', min_value=1, max_value=4, step=1)
instructor_name = st.text_input('Enter Head Instructor Name', '')
class_code = st.text_input('Enter Class Name (All Capitals)', '')

####### IMPORTANT, ENTER INFO HERE ###########
##############################################
#report_number = 1              ##############
#instructor_name = 'Thulith'    ##############
#class_code = 'KX1E18'          ##############
starting_column = 4+4*(report_number) ########
##############################################
##############################################
uploadedTemplate = st.file_uploader('Upload Template', type=['docx'],accept_multiple_files=False,key="Upload template")
uploadedWorkbook = st.file_uploader('Upload Logs', type=['xlsx'],accept_multiple_files=False,key="Upload Student Logs")

#current_directory = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop') ##Create destination folder for reports
#file_destination = os.path.join(f"{current_directory}\{class_code}_Report_{report_number}")
cwd = os.getcwd()
st.write(cwd)

if st.button('Generate Report'):
    #os.makedirs(file_destination)
    workbook = xl.load_workbook(uploadedWorkbook) ######## Generate automated excel workbook ########
    batch = workbook[class_code]
    for row in range(4, batch.max_row):###Iterate the rows, and assign all cell values to variables
        student_name = batch.cell(row, 2)
        week1_log = batch.cell(row, starting_column)
        week2_log = batch.cell(row, starting_column+1)
        week3_log = batch.cell(row, starting_column+2)
        week4_log = batch.cell(row, starting_column+3)
        w1_tut_atnd = batch.cell(row, starting_column+32+(report_number-1)*4)
        w1_lab_atnd = batch.cell(row, starting_column+33+(report_number-1)*4)
        w2_tut_atnd = batch.cell(row, starting_column+34+(report_number-1)*4)
        w2_lab_atnd = batch.cell(row, starting_column+35+(report_number-1)*4)
        w3_tut_atnd = batch.cell(row, starting_column+36+(report_number-1)*4)
        w3_lab_atnd = batch.cell(row, starting_column+37+(report_number-1)*4)
        w4_tut_atnd = batch.cell(row, starting_column+38+(report_number-1)*4)
        w4_lab_atnd = batch.cell(row, starting_column+39+(report_number-1)*4)
        report_comment = batch.cell(row, 33+report_number)
        week_num1 = 4*(report_number)-3
        week_num2 = week_num1+1 
        week_num3 = week_num1+2
        week_num4 = week_num1+3
        template = DocxTemplate(uploadedTemplate) ##Open word template
        context = {                             #Create a dictionary mapping template variables to program variables
            'studentName': student_name.value,
            'w1_log': week1_log.value,
            'w2_log': week2_log.value,
            'w3_log': week3_log.value,
            'w4_log': week4_log.value,
            'w1tut': w1_tut_atnd.value,
            'w1lab': w1_lab_atnd.value,
            'w2tut': w2_tut_atnd.value,
            'w2lab': w2_lab_atnd.value,
            'w3tut': w3_tut_atnd.value,
            'w3lab': w3_lab_atnd.value,
            'w4tut': w4_tut_atnd.value,
            'w4lab': w4_lab_atnd.value,
            'weekNo1': week_num1,
            'weekNo2': week_num2,
            'weekNo3': week_num3,
            'weekNo4': week_num4,
            'reportComments': report_comment.value,
            'reportNo' : report_number,
            'instructor': instructor_name,
            'classCode': class_code,
            }
        template.render(context)
        filename = f'{student_name.value}_report_{report_number}.docx' 
        #filepath = os.path.join(file_destination,filename)
        template.save(filename)
    #convert(f'{cwd}')  #convert to pdf
